import openpyxl
import os
import RPi.GPIO as GPIO
from mfrc522 import SimpleMFRC522
import time

# GPIO Pin Setup
RED_LED = 21
GREEN_LED = 20
BUZZER = 18

# File Path
RFID_FILE = "/home/pi/Desktop/Last_code/rfid_cards.xlsx"

# Setup GPIO Pins
def setup_gpio():
    GPIO.setmode(GPIO.BCM)
    GPIO.setwarnings(False)
    GPIO.setup(RED_LED, GPIO.OUT)
    GPIO.setup(GREEN_LED, GPIO.OUT)
    GPIO.setup(BUZZER, GPIO.OUT)
    GPIO.output(RED_LED, GPIO.HIGH)  # Turn on red LED initially
    GPIO.output(GREEN_LED, GPIO.LOW)
    GPIO.output(BUZZER, GPIO.LOW)

# Load or Create Workbook (Fixing any potential issues)
def load_or_create_workbook(file_path):
    try:
        if os.path.exists(file_path):
            print("File exists, loading workbook...")
            try:
                # Try loading with a fix for possible broken formatting
                workbook = openpyxl.load_workbook(file_path, read_only=False, keep_vba=False)
                return workbook
            except Exception as e:
                print(f"Error loading workbook: {e}")
                print("Rebuilding workbook...")
                # Recreate the workbook if it fails to load
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "RFID Cards"
                sheet.append(["UID", "Assigned ID", "Name"])  # Add headers
                workbook.save(file_path)
                return workbook
        else:
            print("File doesn't exist, creating new workbook...")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "RFID Cards"
            sheet.append(["UID", "Assigned ID", "Name"])  # Add headers
            workbook.save(file_path)
            return workbook
    except Exception as e:
        print(f"Error loading workbook: {e}")
        raise

# Add Red Separator
def add_red_separator(sheet):
    from openpyxl.styles import PatternFill

    last_row = len(sheet["A"])
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Insert red separator line
    sheet.append([None, None, None])
    for cell in sheet[last_row + 1]:
        cell.fill = red_fill

# Assign Card
def assign_card(file_path):
    reader = SimpleMFRC522()
    workbook = load_or_create_workbook(file_path)
    sheet = workbook.active

    print("Scan a card to assign.")
    try:
        uid, _ = reader.read()
        uid = str(uid)

        # Check if UID already exists
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == uid:
                print("This card is already saved.")
                return

        # Get Assigned ID
        assigned_id = input("Enter Assigned ID (numbers only): ")
        while not assigned_id.isdigit():
            print("Invalid input. Please enter numbers only.")
            assigned_id = input("Enter Assigned ID (numbers only): ")

        # Get Name
        name = input("Enter Name (letters and spaces only): ")
        while not all(x.isalpha() or x.isspace() for x in name):
            print("Invalid input. Please enter letters and spaces only.")
            name = input("Enter Name (letters and spaces only): ")

        # Append to sheet
        sheet.append([uid, assigned_id, name])
        workbook.save(file_path)
        print("Card assigned and data saved.")

        # Indicate success with LEDs and buzzer
        GPIO.output(GREEN_LED, GPIO.HIGH)
        GPIO.output(BUZZER, GPIO.HIGH)
        time.sleep(0.5)
        GPIO.output(BUZZER, GPIO.LOW)
        GPIO.output(GREEN_LED, GPIO.LOW)

    except Exception as e:
        print(f"Error during card assignment: {e}")

# Read Card Data
def read_card_data(file_path):
    reader = SimpleMFRC522()
    workbook = load_or_create_workbook(file_path)
    sheet = workbook.active

    print("Scan a card to read its data.")
    try:
        uid, _ = reader.read()
        uid = str(uid)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == uid:
                print(f"UID: {row[0]} | Assigned ID: {row[1]} | Name: {row[2]}")

                # Indicate success with LEDs and buzzer
                GPIO.output(GREEN_LED, GPIO.HIGH)
                GPIO.output(BUZZER, GPIO.HIGH)
                time.sleep(0.5)
                GPIO.output(BUZZER, GPIO.LOW)
                GPIO.output(GREEN_LED, GPIO.LOW)
                return

        print("Card not found in the database.")

        # Indicate failure with red LED
        GPIO.output(RED_LED, GPIO.HIGH)
        time.sleep(0.5)
        GPIO.output(RED_LED, GPIO.LOW)

    except Exception as e:
        print(f"Error during card reading: {e}")

# Delete Card Data
def delete_card(file_path):
    reader = SimpleMFRC522()
    workbook = load_or_create_workbook(file_path)
    sheet = workbook.active

    print("Scan a card to delete its data.")
    try:
        uid, _ = reader.read()
        uid = str(uid)

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == uid:
                sheet.delete_rows(i)
                workbook.save(file_path)
                print("Card data deleted successfully.")

                # Indicate success with LEDs and buzzer
                GPIO.output(RED_LED, GPIO.HIGH)
                GPIO.output(BUZZER, GPIO.HIGH)
                time.sleep(0.5)
                GPIO.output(BUZZER, GPIO.LOW)
                GPIO.output(RED_LED, GPIO.LOW)
                return

        print("Card not found in the database.")

    except Exception as e:
        print(f"Error during card deletion: {e}")

# Main Function
def main():
    setup_gpio()
    try:
        while True:
            print("\n1. Assign Card")
            print("2. Read Card Data")
            print("3. Delete Card Data")
            print("4. Add Red Separator")
            print("5. Exit")

            choice = input("Enter your choice: ")

            if choice == "1":
                assign_card(RFID_FILE)
            elif choice == "2":
                read_card_data(RFID_FILE)
            elif choice == "3":
                delete_card(RFID_FILE)
            elif choice == "4":
                workbook = load_or_create_workbook(RFID_FILE)
                sheet = workbook.active
                add_red_separator(sheet)
                workbook.save(RFID_FILE)
                print("Red separator added.")
            elif choice == "5":
                print("Exiting program.")
                break
            else:
                print("Invalid choice. Please try again.")

    finally:
        GPIO.cleanup()
        print("GPIO cleaned up. Goodbye!")

if __name__ == "__main__":
    main()
